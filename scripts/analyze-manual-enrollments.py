import json
import re
import sys
from datetime import date, datetime

from openpyxl import load_workbook


def normalize_phone(value):
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) == 10 and digits.startswith("1"):
        digits = "0" + digits
    if len(digits) == 11 and digits.startswith("010"):
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    return None


def normalize_birth(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y.%m.%d", "%Y-%m-%d", "%m-%d-%y", "%Y.%m.%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def age_on(birth, on_date):
    return on_date.year - birth.year - ((on_date.month, on_date.day) < (birth.month, birth.day))


def main():
    if len(sys.argv) != 2:
        raise SystemExit("사용법: python3 scripts/analyze-manual-enrollments.py <xlsx-path>")

    sheet = load_workbook(sys.argv[1], data_only=True).active
    headers = [cell.value for cell in sheet[2]]
    entries = [dict(zip(headers, row)) for row in sheet.iter_rows(min_row=3, values_only=True) if any(row)]
    today = date.today()

    email_missing = []
    invalid_phone = []
    birth_missing = []
    birth_unparseable = []
    minors = []
    normalized = []

    for row in entries:
        name = str(row.get("이름") or "").strip()
        email = str(row.get("이메일") or "").strip().lower()
        phone = normalize_phone(row.get("연락처"))
        birth = normalize_birth(row.get("생년월일"))
        if not email:
            email_missing.append(name)
        if not phone:
            invalid_phone.append(name)
        if not birth:
            (birth_missing if not row.get("생년월일") else birth_unparseable).append(name)
        elif age_on(birth, today) < 19:
            minors.append(name)
        normalized.append({"name": name, "phone": phone, "email": email or None, "branch": row.get("지점")})

    summary = {
        "total_rows": len(entries),
        "email_missing_count": len(email_missing),
        "birth_missing_count": len(birth_missing),
        "birth_unparseable_count": len(birth_unparseable),
        "invalid_phone_count": len(invalid_phone),
        "minor_count": len(minors),
        "email_missing_names": email_missing,
        "birth_missing_names": birth_missing,
        "birth_unparseable_names": birth_unparseable,
        "invalid_phone_names": invalid_phone,
        "minor_names": minors,
        "candidate_phone_count": sum(1 for item in normalized if item["phone"]),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
